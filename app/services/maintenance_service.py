"""Rotinas de manutencao para manter dados ativos leves e arquivados."""

from __future__ import annotations

import json
import os
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config.paths import (
    get_backups_dir,
    get_bot_config_db_path,
    get_bot_log_path,
    get_data_dir,
    get_metas_db_path,
    get_ocorrencias_db_path,
    get_pdfs_dir,
    get_ponto_db_path,
    get_tarefas_db_path,
)
from app.config.settings import (
    SHEET_BOT_QUEUE,
    SHEET_GOALS,
    SHEET_OCCURRENCES,
    SHEET_SENT_REMINDERS,
    SHEET_TASK_CHECKS,
    SHEET_TIME_RECORDS,
)
from app.utils.dates import format_datetime, parse_date, parse_datetime


DEFAULT_RETENTION_MONTHS = 3
DEFAULT_MAX_PDFS_PER_FOLDER = 30
DEFAULT_LOG_MAX_MB = 5
DEFAULT_LOG_RETENTION_DAYS = 60


class MaintenanceService:
    def __init__(
        self,
        data_dir: str | None = None,
        pdfs_dir: str | None = None,
        backups_dir: str | None = None,
        log_path: str | None = None,
        today_provider=None,
    ):
        self.data_dir = data_dir or get_data_dir()
        self.pdfs_dir = pdfs_dir or get_pdfs_dir()
        self.backups_dir = backups_dir or get_backups_dir()
        self.log_path = log_path or get_bot_log_path()
        self.today_provider = today_provider or date.today
        self.state_path = os.path.join(self.data_dir, "maintenance_state.json")

    def run(
        self,
        retention_months: int = DEFAULT_RETENTION_MONTHS,
        max_pdfs_per_folder: int = DEFAULT_MAX_PDFS_PER_FOLDER,
    ) -> dict[str, Any]:
        summary = {
            "arquivados": self.archive_old_data(retention_months),
            "pdfs_removidos": self.cleanup_pdfs(max_pdfs_per_folder),
            "logs": self.cleanup_logs(),
            "executado_em": format_datetime(),
        }
        return summary

    def run_auto_if_due(
        self,
        retention_months: int = DEFAULT_RETENTION_MONTHS,
        max_pdfs_per_folder: int = DEFAULT_MAX_PDFS_PER_FOLDER,
    ) -> dict[str, Any]:
        today_text = self.today_provider().isoformat()
        state = self._read_state()
        if state.get("last_run") == today_text:
            return {"executado": False, "motivo": "Manutencao ja executada hoje."}
        summary = self.run(retention_months, max_pdfs_per_folder)
        state["last_run"] = today_text
        state["last_summary"] = summary
        self._write_state(state)
        return {"executado": True, **summary}

    def archive_old_data(self, retention_months: int = DEFAULT_RETENTION_MONTHS) -> dict[str, int]:
        cutoff = _retention_cutoff(self.today_provider(), retention_months)
        specs = [
            (os.path.join(self.data_dir, "ponto.xlsx"), "ponto", {SHEET_TIME_RECORDS: "data"}),
            (os.path.join(self.data_dir, "tarefas_pops.xlsx"), "tarefas_pops", {SHEET_TASK_CHECKS: "data"}),
            (os.path.join(self.data_dir, "ocorrencias.xlsx"), "ocorrencias", {SHEET_OCCURRENCES: "data"}),
            (
                os.path.join(self.data_dir, "bot_config.xlsx"),
                "bot_config",
                {SHEET_SENT_REMINDERS: "data", SHEET_BOT_QUEUE: "data"},
            ),
            (os.path.join(self.data_dir, "metas.xlsx"), "metas", {SHEET_GOALS: "periodo_mes"}),
        ]
        # Mantem compatibilidade com testes e overrides antigos.
        specs = [
            (get_ponto_db_path() if self.data_dir == get_data_dir() else path, stem, sheets)
            if stem == "ponto"
            else (get_tarefas_db_path() if self.data_dir == get_data_dir() else path, stem, sheets)
            if stem == "tarefas_pops"
            else (get_ocorrencias_db_path() if self.data_dir == get_data_dir() else path, stem, sheets)
            if stem == "ocorrencias"
            else (get_bot_config_db_path() if self.data_dir == get_data_dir() else path, stem, sheets)
            if stem == "bot_config"
            else (get_metas_db_path() if self.data_dir == get_data_dir() else path, stem, sheets)
            for path, stem, sheets in specs
        ]

        summary: dict[str, int] = {}
        for workbook_path, stem, sheets in specs:
            if not os.path.isfile(workbook_path):
                continue
            removed = self._archive_workbook(workbook_path, stem, sheets, cutoff)
            if removed:
                summary[stem] = removed
        return summary

    def cleanup_pdfs(self, max_per_folder: int = DEFAULT_MAX_PDFS_PER_FOLDER) -> int:
        if not os.path.isdir(self.pdfs_dir):
            return 0
        max_per_folder = max(int(max_per_folder or DEFAULT_MAX_PDFS_PER_FOLDER), 1)
        removed = 0
        for root, _dirs, files in os.walk(self.pdfs_dir):
            pdfs = [Path(root) / name for name in files if name.lower().endswith(".pdf")]
            excess = len(pdfs) - max_per_folder
            if excess <= 0:
                continue
            pdfs.sort(key=lambda item: item.stat().st_mtime if item.exists() else 0)
            for path in pdfs[:excess]:
                try:
                    path.unlink()
                    removed += 1
                except OSError:
                    pass
        return removed

    def cleanup_logs(
        self,
        max_mb: int = DEFAULT_LOG_MAX_MB,
        keep_days: int = DEFAULT_LOG_RETENTION_DAYS,
    ) -> dict[str, int]:
        summary = {"rotacionados": 0, "removidos": 0}
        log_path = Path(self.log_path)
        if log_path.is_file() and log_path.stat().st_size > max(max_mb, 1) * 1024 * 1024:
            folder = Path(self.backups_dir) / "logs"
            folder.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = folder / f"{log_path.stem}_{stamp}{log_path.suffix}"
            try:
                shutil.move(str(log_path), str(dest))
                summary["rotacionados"] += 1
            except OSError:
                pass

        logs_dir = Path(self.backups_dir) / "logs"
        if logs_dir.is_dir():
            cutoff = datetime.now() - timedelta(days=max(int(keep_days or DEFAULT_LOG_RETENTION_DAYS), 1))
            for path in logs_dir.glob("*.log"):
                try:
                    if datetime.fromtimestamp(path.stat().st_mtime) < cutoff:
                        path.unlink()
                        summary["removidos"] += 1
                except OSError:
                    pass
        return summary

    def _archive_workbook(self, workbook_path: str, stem: str, sheets: dict[str, str], cutoff: date) -> int:
        wb = load_workbook(workbook_path)
        changed = False
        total_removed = 0
        try:
            for sheet_name, date_field in sheets.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                headers = _headers(ws)
                if not headers or date_field not in headers:
                    continue
                keep_rows: list[dict[str, Any]] = []
                archive_by_month: dict[str, list[dict[str, Any]]] = {}
                for row in _sheet_rows(ws, headers):
                    row_date = _row_date(row.get(date_field), month_field=date_field == "periodo_mes")
                    if row_date and row_date < cutoff:
                        month_key = row_date.strftime("%Y-%m")
                        archive_by_month.setdefault(month_key, []).append(row)
                    else:
                        keep_rows.append(row)
                if not archive_by_month:
                    continue
                for month_key, rows in archive_by_month.items():
                    self._append_archive(month_key, stem, sheet_name, headers, rows)
                    total_removed += len(rows)
                _rewrite_sheet(ws, headers, keep_rows)
                changed = True
            if changed:
                wb.save(workbook_path)
        finally:
            wb.close()
        return total_removed

    def _append_archive(self, month_key: str, stem: str, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
        folder = Path(self.backups_dir) / "arquivos_mensais" / month_key
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / f"{stem}_{month_key}.xlsx"
        if path.is_file():
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
            else:
                ws = wb[sheet_name]
                if ws.max_row < 1:
                    ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(headers)
        try:
            for row in rows:
                ws.append([row.get(header, "") for header in headers])
            wb.save(path)
        finally:
            wb.close()

    def _read_state(self) -> dict[str, Any]:
        try:
            with open(self.state_path, "r", encoding="utf-8") as file:
                return json.load(file)
        except (OSError, ValueError):
            return {}

    def _write_state(self, state: dict[str, Any]) -> None:
        os.makedirs(os.path.dirname(self.state_path), exist_ok=True)
        with open(self.state_path, "w", encoding="utf-8") as file:
            json.dump(state, file, ensure_ascii=False, indent=2)


def _retention_cutoff(today: date, retention_months: int) -> date:
    retention_months = max(int(retention_months or DEFAULT_RETENTION_MONTHS), 1)
    year = today.year
    month = today.month - (retention_months - 1)
    while month <= 0:
        month += 12
        year -= 1
    return date(year, month, 1)


def _headers(ws: Worksheet) -> list[str]:
    return [str(cell.value).strip() for cell in ws[1] if cell.value is not None]


def _sheet_rows(ws: Worksheet, headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if values is None or all(value is None for value in values):
            continue
        rows.append({headers[index]: value for index, value in enumerate(values)})
    return rows


def _rewrite_sheet(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def _row_date(value: Any, month_field: bool = False) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    if month_field:
        for fmt in ("%m/%Y", "%Y-%m"):
            try:
                parsed = datetime.strptime(text, fmt)
                return date(parsed.year, parsed.month, 1)
            except ValueError:
                pass
    for parser in (parse_date, parse_datetime):
        try:
            parsed = parser(text)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception:
            pass
    return None
