from __future__ import annotations

import os
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.repositories.excel_database import ExcelDatabase
from app.repositories.excel_schema import POINT_SHEETS_CONFIG
from app.config.settings import SHEET_TIME_RECORDS
from app.services.maintenance_service import MaintenanceService


def test_maintenance_archives_old_rows_and_keeps_recent_rows(tmp_path):
    data_dir = tmp_path / "data"
    backups_dir = tmp_path / "backups"
    pdfs_dir = tmp_path / "pdfs"
    db = ExcelDatabase(
        db_path=str(data_dir / "ponto.xlsx"),
        sheets_config=POINT_SHEETS_CONFIG,
        backup_dir=str(backups_dir),
        backup_stem="ponto",
    )
    db.ensure_database()
    db.write_sheet(
        SHEET_TIME_RECORDS,
        POINT_SHEETS_CONFIG[SHEET_TIME_RECORDS],
        [
            {
                "ponto_id": "old",
                "colaborador_id": "1",
                "nome_colaborador": "Antigo",
                "tipo_ponto": "entrada",
                "data": "15/02/2026",
                "hora": "08:00",
                "data_hora": "15/02/2026 08:00:00",
                "observacoes": "",
            },
            {
                "ponto_id": "new",
                "colaborador_id": "1",
                "nome_colaborador": "Recente",
                "tipo_ponto": "entrada",
                "data": "10/05/2026",
                "hora": "08:00",
                "data_hora": "10/05/2026 08:00:00",
                "observacoes": "",
            },
        ],
    )
    service = MaintenanceService(
        data_dir=str(data_dir),
        pdfs_dir=str(pdfs_dir),
        backups_dir=str(backups_dir),
        today_provider=lambda: date(2026, 5, 21),
    )

    summary = service.archive_old_data(retention_months=3)

    assert summary == {"ponto": 1}
    active_rows = db.read_sheet(SHEET_TIME_RECORDS)
    assert [row["ponto_id"] for row in active_rows] == ["new"]
    archive = backups_dir / "arquivos_mensais" / "2026-02" / "ponto_2026-02.xlsx"
    assert archive.exists()
    wb = load_workbook(archive, read_only=True)
    try:
        rows = list(wb[SHEET_TIME_RECORDS].iter_rows(min_row=2, values_only=True))
        assert rows[0][0] == "old"
    finally:
        wb.close()


def test_maintenance_pdf_cleanup_keeps_newest_files(tmp_path):
    pdf_dir = tmp_path / "pdfs" / "relatorios"
    pdf_dir.mkdir(parents=True)
    paths = [pdf_dir / f"relatorio_{idx}.pdf" for idx in range(3)]
    for idx, path in enumerate(paths):
        path.write_text("pdf", encoding="utf-8")
        os.utime(path, (idx + 1, idx + 1))
    service = MaintenanceService(data_dir=str(tmp_path / "data"), pdfs_dir=str(tmp_path / "pdfs"), backups_dir=str(tmp_path / "backups"))

    removed = service.cleanup_pdfs(max_per_folder=2)

    assert removed == 1
    assert not paths[0].exists()
    assert paths[1].exists()
    assert paths[2].exists()


def test_maintenance_rotates_large_bot_log(tmp_path):
    data_dir = tmp_path / "data"
    backups_dir = tmp_path / "backups"
    log_path = data_dir / "bot_whatsapp.log"
    data_dir.mkdir()
    log_path.write_bytes(b"x" * (2 * 1024 * 1024))
    service = MaintenanceService(
        data_dir=str(data_dir),
        pdfs_dir=str(tmp_path / "pdfs"),
        backups_dir=str(backups_dir),
        log_path=str(log_path),
    )

    summary = service.cleanup_logs(max_mb=1)

    assert summary["rotacionados"] == 1
    assert not log_path.exists()
    assert list((backups_dir / "logs").glob("bot_whatsapp_*.log"))
