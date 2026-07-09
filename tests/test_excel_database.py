from __future__ import annotations

import json
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook

from app.config.paths import (
    get_colaboradores_db_path,
    get_data_dir,
    get_db_version_path,
    set_root_override,
)
from app.repositories.excel_database import ExcelDatabase
from app.repositories.excel_schema import COLLABORATOR_HEADERS, JOURNEY_HEADERS, OCCURRENCE_HEADERS, SECTOR_HEADERS

from app.config.settings import (
    OCCURRENCE_TASK_MISSED,
    SHEET_COLLABORATORS,
    SHEET_OCCURRENCES,
    SHEET_TASK_CHECKS,
    SHEET_TASKS,
    SHEET_TIME_RECORDS,
    SHEET_JOURNEYS,
    SHEET_SECTORS,
)


def test_excel_databases_are_created_with_expected_sheets(stack):
    dbs = stack["dbs"]

    assert dbs["colaboradores"].sheet_names() == [SHEET_COLLABORATORS]
    assert dbs["ponto"].sheet_names() == [SHEET_TIME_RECORDS, SHEET_JOURNEYS]
    assert dbs["tarefas"].sheet_names() == [SHEET_TASKS, SHEET_TASK_CHECKS]
    assert dbs["ocorrencias"].sheet_names() == [SHEET_OCCURRENCES]
    assert dbs["setores"].sheet_names() == [SHEET_SECTORS]


def test_separated_workbook_data_flow(stack):
    collaborator = stack["collaborators"].create_collaborator("Ana Souza", setor="Atendimento")
    task = stack["tasks"].create_task("Limpar loja", nome_setor="Todos", horario_inicio="12:00", horario_limite="13:00")

    stack["time"].record_time(collaborator["colaborador_id"], "entrada", when=datetime(2026, 5, 4, 8, 0))
    stack["tasks"].mark_done(task["tarefa_id"], collaborator["colaborador_id"], when=datetime(2026, 5, 4, 12, 30))
    occurrence = stack["occurrences"].create_occurrence(
        tipo=OCCURRENCE_TASK_MISSED,
        descricao="Teste de persistencia separada.",
        tarefa_id=task["tarefa_id"],
        nome_tarefa=task["nome"],
    )

    assert stack["dbs"]["colaboradores"].read_sheet(SHEET_COLLABORATORS)
    assert stack["dbs"]["ponto"].read_sheet(SHEET_TIME_RECORDS)
    assert stack["dbs"]["tarefas"].read_sheet(SHEET_TASKS)
    assert stack["dbs"]["tarefas"].read_sheet(SHEET_TASK_CHECKS)
    saved_occurrences = stack["dbs"]["ocorrencias"].read_sheet(SHEET_OCCURRENCES)
    assert len(saved_occurrences) == 1
    assert saved_occurrences[0]["ocorrencia_id"] == occurrence["ocorrencia_id"]
    assert saved_occurrences[0]["tipo"] == OCCURRENCE_TASK_MISSED


def test_excel_headers_are_optimized(stack):
    dbs = stack["dbs"]
    colaboradores = dbs["colaboradores"].load_workbook_safe()
    ponto = dbs["ponto"].load_workbook_safe()
    ocorrencias = dbs["ocorrencias"].load_workbook_safe()
    setores = dbs["setores"].load_workbook_safe()
    try:
        assert [cell.value for cell in colaboradores[SHEET_COLLABORATORS][1]] == COLLABORATOR_HEADERS
        assert [cell.value for cell in ponto[SHEET_JOURNEYS][1]] == JOURNEY_HEADERS
        assert [cell.value for cell in ocorrencias[SHEET_OCCURRENCES][1]] == OCCURRENCE_HEADERS
        assert [cell.value for cell in setores[SHEET_SECTORS][1]] == SECTOR_HEADERS
        assert "Jornada" not in ponto.sheetnames
        assert "tipo_jornada" in JOURNEY_HEADERS
        assert "tempo_intervalo" in JOURNEY_HEADERS
        assert "pausa_inicio" not in JOURNEY_HEADERS
        assert "pausa_fim" not in JOURNEY_HEADERS
        assert "tipo_escala" not in JOURNEY_HEADERS
    finally:
        colaboradores.close()
        ponto.close()
        ocorrencias.close()
        setores.close()


def test_schema_migration_adds_columns_without_removing_existing_data(tmp_path):
    data_dir = tmp_path / "data"
    backups_dir = tmp_path / "backups"
    db_path = data_dir / "colaboradores.xlsx"
    data_dir.mkdir()
    old_headers = [header for header in COLLABORATOR_HEADERS if header != "observacoes"] + ["campo_antigo"]
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_COLLABORATORS
    ws.append(old_headers)
    ws.append(
        [
            "colab-1",
            "Ana",
            "Operadora",
            "32999990000",
            "",
            "Atendimento",
            1200,
            "",
            50,
            75,
            "ativo",
            "01/06/2026",
            "01/06/2026 08:00:00",
            "valor legado",
        ]
    )
    wb.save(db_path)
    wb.close()
    other_wb = Workbook()
    other_wb.active.title = "Outra"
    other_wb.save(data_dir / "tarefas_pops.xlsx")
    other_wb.close()

    db = ExcelDatabase(
        db_path=str(db_path),
        sheets_config={SHEET_COLLABORATORS: COLLABORATOR_HEADERS},
        backup_dir=str(backups_dir),
        backup_stem="colaboradores",
    )

    db.ensure_database()

    migrated = load_workbook(db_path)
    try:
        headers = [cell.value for cell in migrated[SHEET_COLLABORATORS][1]]
        row = [cell.value for cell in migrated[SHEET_COLLABORATORS][2]]
    finally:
        migrated.close()

    assert "observacoes" in headers
    assert "campo_antigo" in headers
    assert row[headers.index("nome")] == "Ana"
    assert row[headers.index("campo_antigo")] == "valor legado"
    backup_folders = list(backups_dir.glob("backup_*"))
    assert backup_folders
    assert (backup_folders[0] / "colaboradores.xlsx").exists()
    assert (backup_folders[0] / "tarefas_pops.xlsx").exists()


def test_db_version_file_is_written_for_real_app_data_workbook(tmp_path):
    set_root_override(str(tmp_path))
    try:
        db = ExcelDatabase(
            db_path=get_colaboradores_db_path(),
            sheets_config={SHEET_COLLABORATORS: COLLABORATOR_HEADERS},
            backup_stem="colaboradores",
        )

        db.ensure_database()

        version_path = get_db_version_path()
        with open(version_path, "r", encoding="utf-8") as file:
            version = json.load(file)
        assert version["version"]
        assert version["last_migration"]
    finally:
        set_root_override(None)


def test_packaged_paths_use_executable_folder(monkeypatch, tmp_path):
    set_root_override(None)
    fake_exe = tmp_path / "ControlePontoTarefas.exe"
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(fake_exe))

    data_dir = get_data_dir()

    assert data_dir == str(tmp_path / "data")
